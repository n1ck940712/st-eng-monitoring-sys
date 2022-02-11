from modules import database, logger
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl import Workbook
from pathlib import Path
import os, subprocess

# return definition
# 0: ok
# 1: error
# 2: no usb drive found

def export_usb(log, files=[]):
    try:
        os.putenv('FOO', ' '.join(files))
        result = subprocess.Popen("bash /home/pi/monitoring_sys/python/export_usb.sh", shell=True, stdout=subprocess.PIPE).stdout.read().decode('utf-8')
        if result == '0':
            log.info('%s report(s) exported' % len(files))
            return 0    
        elif result == '1':
            log.error('no USB driver found')
            return 2
        else:
            return 1
    except Exception as e:
        log.error('Export USB ERROR. %s' % e)
        return 1
    
def draw_box(ws, start_x, end_x, start_y, end_y):
    thin = Side(border_style="thin", color="000000")
    width = end_x - start_x
    height = end_y - start_y
    if (width > 1) and (height > 1):
        # corners
        ws.cell(row=start_y, column=start_x).border = Border(top=thin, left=thin)
        ws.cell(row=end_y, column=start_x).border = Border(bottom=thin, left=thin)
        ws.cell(row=start_y, column=end_x).border = Border(top=thin, right=thin)
        ws.cell(row=end_y, column=end_x).border = Border(bottom=thin, right=thin)
        # # horizontal line
        for i in range(width-1): 
            ws.cell(row=start_y, column=start_x+1+i).border = Border(top=thin)
            ws.cell(row=end_y, column=start_x+1+i).border = Border(bottom=thin)
        # # vertical line
        for i in range(height-1):
            ws.cell(row=start_y+1+i, column=start_x).border = Border(left=thin)
            ws.cell(row=start_y+1+i, column=end_x).border = Border(right=thin)
    elif (width == 0) and (height > 1):
        ws.cell(row=start_y, column=start_x).border = Border(top=thin, left=thin, right=thin)
        ws.cell(row=end_y, column=end_x).border = Border(bottom=thin, left=thin, right=thin)
        # # vertical line
        for i in range(height-1):
            ws.cell(row=start_y+1+i, column=start_x).border = Border(left=thin, right=thin)
    elif (height == 0) and (width > 1):
        ws.cell(row=start_y, column=start_x).border = Border(top=thin, left=thin, bottom=thin)
        ws.cell(row=end_y, column=end_x).border = Border(top=thin, right=thin, bottom=thin)
        # # horizontal line
        for i in range(width-1): 
            ws.cell(row=start_y, column=start_x+1+i).border = Border(top=thin, bottom=thin)
    elif (height == 1) and (width == 1):
        # corners
        ws.cell(row=start_y, column=start_x).border = Border(top=thin, left=thin)
        ws.cell(row=end_y, column=start_x).border = Border(bottom=thin, left=thin)
        ws.cell(row=start_y, column=end_x).border = Border(top=thin, right=thin)
        ws.cell(row=end_y, column=end_x).border = Border(bottom=thin, right=thin)
    elif (height == 0) and (width == 1):
        ws.cell(row=start_y, column=start_x).border = Border(top=thin, left=thin, bottom=thin)
        ws.cell(row=end_y, column=end_x).border = Border(top=thin, right=thin, bottom=thin)
    elif (width == 0) and (height == 1):
        ws.cell(row=start_y, column=start_x).border = Border(top=thin, left=thin, right=thin)
        ws.cell(row=end_y, column=end_x).border = Border(bottom=thin, left=thin, right=thin)
    elif (width == 0) and (height == 0):
        ws.cell(row=start_y, column=start_x).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    else:
        print('invalid')
    return

def generate(log, batch_id):
    thin = Side(border_style="thin", color="000000")
    report = database.query('local', 'get', 'SELECT type, time_completed FROM app_report WHERE id="%s"' % batch_id)
    if len(report) > 0: 
        process_name = report[0][0]
        process_date = report[0][1]
        process_date_input = process_date.strftime("%-d %b %Y")

        column_num = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z','aa','ab','ac','ad','ae','af','ag','ah','ai','aj','ak','al','am','an','ao','ap','aq','ar','as','at','au','av','aw','ax','ay','az',]
        wb = Workbook()
        ws = wb.active
        for i in column_num:
            ws.column_dimensions[i].width = 1.724137931034483
        
        font_size_10 = Font(size=10,)
        font_size_10_underline = Font(size=10, underline='single')
        font_size_11_bold = Font(size=11, bold=True)
        
        if process_name == 'pre production': 
            row = database.query('local', 'get', 'SELECT flow, timestamp FROM app_logdata WHERE batch="%s" ORDER BY timestamp DESC' % batch_id)
            if len(row) == 2:
                valid_data = True
                start_purge_time = row[1][1].strftime("%H:%M:%S")
                start_n2_flow_rate = row[1][0]
                end_purge_time = row[0][1].strftime("%H:%M:%S")
                end_n2_flow_rate = row[0][0]
                report_name = "pre_production_%s" % (process_date.strftime("%-d%b%Y"))

                ws.merge_cells(start_row=6, start_column=6, end_row=6, end_column=13)
                ws.merge_cells(start_row=19, start_column=12, end_row=19, end_column=15)
                ws.merge_cells(start_row=19, start_column=38, end_row=19, end_column=41)
                ws.merge_cells(start_row=18, start_column=6, end_row=18, end_column=11)
                ws.merge_cells(start_row=18, start_column=32, end_row=18, end_column=37)
                
                ws['AH1'] = 'FORM 382-130001-Z-PROC-02 REV 01'
                ws['AH1'].font = font_size_10
                ws['U2'] = 'PRODUCTION FORM'
                ws['U2'].font = font_size_11_bold
                ws['C4'] = 'Material:'
                ws['C4'].font = font_size_10
                ws['H4'] = 'HTPB'
                ws['H4'].font = font_size_10
                ws['C5'] = 'Lot number/Drum number:'
                ws['C5'].font = font_size_10
                ws['C6'] = 'Date:'
                ws['C6'].font = font_size_10
                ws['F6'] = process_date.strftime("%-d %b %Y")
                ws['F6'].font = font_size_10

                ws['C9'] = 'Nitrogen gas supply:'
                ws['C9'].font = font_size_11_bold
                ws['C10'] = 'Pre-operation check'
                ws['C10'].font = font_size_10_underline
                ws['C11'] = 'Pressure of nitrogen pallet in use:'
                ws['C11'].font = font_size_10
                ws['S11'] = '______'
                ws['S11'].font = font_size_10
                ws['W11'] = 'bar'
                ws['W11'].font = font_size_10
                ws['C12'] = 'If less than 120 bar,'
                ws['C12'].font = font_size_10
                ws['C13'] = 'Pressure of second nitrogen pallet:'
                ws['C13'].font = font_size_10
                ws['T13'] = '______'
                ws['T13'].font = font_size_10
                ws['X13'] = 'bar'
                ws['X13'].font = font_size_10
                ws['AC10'] = 'After operation check'
                ws['AC10'].font = font_size_10_underline
                ws['AC11'] = 'Pressure of nitrogen pallet in use:'
                ws['AC11'].font = font_size_10
                ws['AS11'] = '______'
                ws['AS11'].font = font_size_10
                ws['AW11'] = 'bar'
                ws['AW11'].font = font_size_10
                ws['AC12'] = 'If less than 120 bar,'
                ws['AC12'].font = font_size_10
                ws['AC13'] = 'Pressure of second nitrogen pallet:'
                ws['AC13'].font = font_size_10
                ws['AT13'] = '______'
                ws['AT13'].font = font_size_10
                ws['AW13'] = 'bar'
                ws['AW13'].font = font_size_10

                ws['C16'] = 'Transferring'
                ws['C16'].font = font_size_11_bold
                ws['J16'] = '________'
                ws['J16'].font = font_size_10
                ws['O16'] = 'to vessel - Purging of Nitrogen for 5 minutes'
                ws['O16'].font = font_size_11_bold
                ws['C17'] = 'Start'
                ws['C17'].font = font_size_10_underline
                ws['C18'] = 'Time:'
                ws['C18'].font = font_size_10
                ws['F18'] = start_purge_time
                ws['F18'].font = font_size_10_underline
                ws['C19'] = 'Nitrogen flow rate:'
                ws['C19'].font = font_size_10
                ws['L19'] = start_n2_flow_rate
                ws['L19'].font = font_size_10_underline
                ws['P19'] = 'L/min'
                ws['P19'].font = font_size_10
                ws['AC17'] = 'End'
                ws['AC17'].font = font_size_10_underline
                ws['AC18'] = 'Time:'
                ws['AC18'].font = font_size_10
                ws['AF18'] = end_purge_time
                ws['AF18'].font = font_size_10_underline
                ws['AC19'] = 'Nitrogen flow rate:'
                ws['AC19'].font = font_size_10
                ws['AL19'] = end_n2_flow_rate
                ws['AL19'].font = font_size_10_underline
                ws['AP19'] = 'L/min'
                ws['AP19'].font = font_size_10

                ws['C22'] = 'Vacuum Pump'
                ws['C22'].font = font_size_11_bold
                ws['C23'] = 'Vacuum pump used: Vacuum pump'
                ws['C23'].font = font_size_10
                ws['T23'] = '_______________'
                ws['T23'].font = font_size_10

                ws['C26'] = 'Weight of HTPB, Treated'
                ws['C26'].font = font_size_11_bold
                ws['C27'] = 'Weighing Balance S/No:'
                ws['C27'].font = font_size_10
                ws['O27'] = '____________'
                ws['O27'].font = font_size_10

                ws['AC29'] = 'Weight (kg)'
                ws['AC29'].font = font_size_11_bold
                ws['C30'] = 'Empty 220-L drum'
                ws['C30'].font = font_size_10
                ws['C31'] = 'Filled 220-L drum with HTPB, Treated'
                ws['C31'].font = font_size_10
                ws['C32'] = 'Amount of HTPB, Treated collected'
                ws['C32'].font = font_size_10

                ws['C36'] = 'Done by:'
                ws['C36'].font = font_size_10
                ws['C40'] = 'Checked by:'
                ws['C40'].font = font_size_10
                
                draw_box(ws, 2, 51, 2, 51)
                draw_box(ws, 2, 51, 2, 7)
                draw_box(ws, 2, 51, 8, 14)
                draw_box(ws, 2, 51, 15, 20)
                draw_box(ws, 2, 51, 21, 24)

                draw_box(ws, 3, 39, 29, 32)
                draw_box(ws, 23, 39, 29, 32)

                draw_box(ws, 3, 22, 30, 30)
                draw_box(ws, 23, 39, 30, 30)

                draw_box(ws, 3, 22, 32, 32)
                draw_box(ws, 23, 39, 32, 32)

            else:
                log.error('invalid log data')
                valid_data = False

        elif process_name == 'production': 
            rows = database.query('local', 'get', 'SELECT type, flow, flow_unit, temperature, temperature_unit, heater_set, heater_set_unit, pressure, pressure_unit, timestamp FROM app_logdata WHERE batch="%s" ORDER BY timestamp ASC' % batch_id)
            if len(rows) > 0:
                valid_data = True
                page_number = 1
                table_row_num = 14

                ws.merge_cells(start_row=6, start_column=6, end_row=6, end_column=13) # date 
                ws.merge_cells(start_row=10, start_column=3, end_row=14, end_column=5) # STAGE
                ws['C10'] = 'Stage'
                ws['C10'].alignment = Alignment(vertical='center', horizontal='center')
                ws.merge_cells(start_row=10, start_column=6, end_row=14, end_column=10) # time
                ws['F10'] = 'Time'
                ws['F10'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                ws.merge_cells(start_row=10, start_column=11, end_row=14, end_column=18) # vacuum
                ws['K10'] = 'Vacuum Pressure (mbar) (S/No.: __________)'
                ws['K10'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                ws.merge_cells(start_row=10, start_column=19, end_row=14, end_column=26) # vessel temp
                ws['S10'] = 'Vessel Temp (°C) (S/No.: __________)'
                ws['S10'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                ws.merge_cells(start_row=10, start_column=27, end_row=14, end_column=34) # heater set
                ws['AA10'] = 'Heater Set Point (°C) (S/No.: __________)'
                ws['AA10'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                ws.merge_cells(start_row=10, start_column=35, end_row=14, end_column=42) # nitrogen flow
                ws['AI10'] = 'Nitrogen Flow Rate (L/min) (S/No.: __________)'
                ws['AI10'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                ws.merge_cells(start_row=10, start_column=43, end_row=14, end_column=50) # nitrogen flow
                ws['AQ10'] = 'Notes'
                ws['AQ10'].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

                # row 1
                for i in range (45):
                    if i < 15:
                        pass
                    else:
                        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5) # col 1
                        ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=10) # col 2
                        ws.merge_cells(start_row=i, start_column=11, end_row=i, end_column=18) # col 3
                        ws.merge_cells(start_row=i, start_column=19, end_row=i, end_column=26) # col 4
                        ws.merge_cells(start_row=i, start_column=27, end_row=i, end_column=34) # col 5
                        ws.merge_cells(start_row=i, start_column=35, end_row=i, end_column=42) # col 6
                        ws.merge_cells(start_row=i, start_column=43, end_row=i, end_column=50) # col 7

                for y in range(45):
                    if y < 10:
                        pass
                    else :
                        for x in range(51):
                            if x < 3:
                                pass
                            else:
                                ws.cell(row=y, column=x).border = Border(top=thin, left=thin, bottom=thin, right=thin)
                
                ws['AH1'] = 'FORM 382-130001-Z-PROC-02 REV 01'
                ws['AH1'].font = font_size_10
                ws['U2'] = 'PRODUCTION FORM'
                ws['U2'].font = font_size_11_bold
                ws['C4'] = 'Material:'
                ws['C4'].font = font_size_10
                ws['H4'] = 'HTPB'
                ws['H4'].font = font_size_10
                ws['C5'] = 'Lot number/Drum number:'
                ws['C5'].font = font_size_10
                ws['C6'] = 'Date:'
                ws['C6'].font = font_size_10
                ws['F6'] = process_date.strftime("%-d %b %Y")
                ws['F6'].font = font_size_10

                ws['C9'] = 'Purification Process'
                ws['C9'].font = font_size_11_bold

                ws['C46'] = 'Done by:'
                ws['C46'].font = font_size_10
                ws['C49'] = 'Checked by:'
                ws['C49'].font = font_size_10

                
                draw_box(ws, 2, 51, 2, 51)
                draw_box(ws, 2, 51, 2, 7)
                
                for row in rows:
                    table_row_num += 1
                    process_type = row[0].replace('production ', '')

                    if process_type == 'part 1':
                        process_type = '1'
                    elif process_type == 'part 2':
                        process_type = '1'
                    elif process_type == 'part 3':
                        process_type = '2'
                    
                    flow = row[1]
                    flow_unit = row[2]
                    temperature = row[3]
                    temperature_unit = row[4]
                    heater_set = row[5]
                    heater_set_unit = row[6]
                    pressure = row[7]
                    pressure_unit = row[8]
                    timestamp = row[9].strftime("%H:%M:%S")


                    ws.cell(row=table_row_num, column=3).value = process_type
                    ws.cell(row=table_row_num, column=3).alignment = Alignment(horizontal='center')
                    ws.cell(row=table_row_num, column=6).value = timestamp
                    ws.cell(row=table_row_num, column=6).alignment = Alignment(horizontal='center')
                    ws.cell(row=table_row_num, column=11).value = pressure
                    ws.cell(row=table_row_num, column=11).alignment = Alignment(horizontal='center')
                    ws.cell(row=table_row_num, column=19).value = temperature
                    ws.cell(row=table_row_num, column=19).alignment = Alignment(horizontal='center')
                    ws.cell(row=table_row_num, column=27).value = heater_set
                    ws.cell(row=table_row_num, column=27).alignment = Alignment(horizontal='center')
                    ws.cell(row=table_row_num, column=35).value = flow
                    ws.cell(row=table_row_num, column=35).alignment = Alignment(horizontal='center')
                    
                    
                    if table_row_num > 45:
                        # new page
                        pass
                        
                report_name = "production_%s" % (process_date.strftime("%-d%b%Y"))
            else:
                valid_data = False
                log.error('invalid log data')

        elif process_name == 'post production': 
            day_1_start_datetime = database.query('local', 'get', 'SELECT timestamp FROM app_logdata WHERE batch="%s" AND type="post production day 1" ORDER BY timestamp ASC LIMIT 1' % batch_id)[0][0]
            day_1_end_datetime = database.query('local', 'get', 'SELECT timestamp FROM app_logdata WHERE batch="%s" AND type="post production day 1" ORDER BY timestamp DESC LIMIT 1' % batch_id)[0][0]
            day_2_start_datetime = database.query('local', 'get', 'SELECT timestamp FROM app_logdata WHERE batch="%s" AND type="post production day 2" ORDER BY timestamp ASC LIMIT 1' % batch_id)[0][0]
            day_2_end_datetime = database.query('local', 'get', 'SELECT timestamp FROM app_logdata WHERE batch="%s" AND type="post production day 2" ORDER BY timestamp DESC LIMIT 1' % batch_id)[0][0]
            day_3_start_datetime = database.query('local', 'get', 'SELECT timestamp FROM app_logdata WHERE batch="%s" AND type="post production day 3" ORDER BY timestamp ASC LIMIT 1' % batch_id)[0][0]
            day_3_end_datetime = database.query('local', 'get', 'SELECT timestamp FROM app_logdata WHERE batch="%s" AND type="post production day 3" ORDER BY timestamp DESC LIMIT 1' % batch_id)[0][0]
            rows = database.query('local', 'get', 'SELECT type, temperature, temperature_unit, heater_set, heater_set_unit, timestamp FROM app_logdata WHERE batch="%s" ORDER BY timestamp ASC' % batch_id)
            if len(rows) > 0:
                valid_data = True
                page_number = 1

                draw_box(ws, 2, 51, 2, 51)
                draw_box(ws, 2, 51, 4, 15)
                draw_box(ws, 2, 51, 16, 27)
                draw_box(ws, 2, 51, 28, 39)
                
                ws['AH1'] = 'FORM 382-130001-Z-PROC-01 REV 00'
                ws['AH1'].font = font_size_10
                ws['U2'] = 'CLEANING FORM'
                ws['U2'].font = font_size_11_bold
                # day 1 cleaning
                ws['C5'] = 'Day 1 Cleaning'
                ws['C5'].font = font_size_11_bold
                ws['C6'] = 'Date:'
                ws['C6'].font = font_size_10
                ws['F6'] = day_1_start_datetime.strftime('%-d %b %Y')
                ws['F6'].font = font_size_10_underline
                ws['C7'] = 'Temp set point'
                ws['C7'].font = font_size_10
                ws['T7'] = 'Start time:'
                ws['T7'].font = font_size_10
                ws['Z7'] = day_1_start_datetime.strftime("%H:%M:%S")
                ws['Z7'].font = font_size_10_underline
                ws['AI7'] = 'End time:'
                ws['AI7'].font = font_size_10
                ws['AN7'] = day_1_end_datetime.strftime("%H:%M:%S")
                ws['AN7'].font = font_size_10_underline
                # day 2 cleaning
                ws['C17'] = 'Day 2 Cleaning'
                ws['C17'].font = font_size_11_bold
                ws['C18'] = 'Date:'
                ws['C18'].font = font_size_10
                ws['F18'] = day_2_start_datetime.strftime('%-d %b %Y')
                ws['F18'].font = font_size_10_underline
                ws['C19'] = 'Temp set point'
                ws['C19'].font = font_size_10
                ws['T19'] = 'Start time:'
                ws['T19'].font = font_size_10
                ws['Z19'] = day_2_start_datetime.strftime("%H:%M:%S")
                ws['Z19'].font = font_size_10_underline
                ws['AI19'] = 'End time:'
                ws['AI19'].font = font_size_10
                ws['AN19'] = day_2_end_datetime.strftime("%H:%M:%S")
                ws['AN19'].font = font_size_10_underline
                # day 3 dryign
                ws['C29'] = 'Day 3 Drying'
                ws['C29'].font = font_size_11_bold
                ws['C30'] = 'Date:'
                ws['C30'].font = font_size_10
                ws['F30'] = day_3_start_datetime.strftime('%-d %b %Y')
                ws['F30'].font = font_size_10_underline
                ws['C31'] = 'Temp set point'
                ws['C31'].font = font_size_10
                ws['T31'] = 'Start time:'
                ws['T31'].font = font_size_10
                ws['Z31'] = day_3_start_datetime.strftime("%H:%M:%S")
                ws['Z31'].font = font_size_10_underline
                ws['AI31'] = 'End time:'
                ws['AI31'].font = font_size_10
                ws['AN31'] = day_3_end_datetime.strftime("%H:%M:%S")
                ws['AN31'].font = font_size_10_underline

                ws['C9'] = 'Time'
                ws['C9'].font = font_size_10
                ws['C9'].alignment = Alignment(horizontal='center')
                ws['M9'] = 'Vessel Temp (°C)'
                ws['M9'].font = font_size_10
                ws['M9'].alignment = Alignment(horizontal='center')
                ws['C21'] = 'Time'
                ws['C21'].font = font_size_10
                ws['C21'].alignment = Alignment(horizontal='center')
                ws['M21'] = 'Vessel Temp (°C)'
                ws['M21'].font = font_size_10
                ws['M21'].alignment = Alignment(horizontal='center')
                ws['C33'] = 'Time'
                ws['C33'].font = font_size_10
                ws['C33'].alignment = Alignment(horizontal='center')
                ws['M33'] = 'Vessel Temp (°C)'
                ws['M33'].font = font_size_10
                ws['M33'].alignment = Alignment(horizontal='center')

                ws['C41'] = 'Done by:'
                ws['C41'].font = font_size_10
                ws['C45'] = 'Checked by:'
                ws['C45'].font = font_size_10
                
                ws.merge_cells(start_row=6, start_column=6, end_row=6, end_column=11)
                ws.merge_cells(start_row=7, start_column=26, end_row=7, end_column=31)
                ws.merge_cells(start_row=7, start_column=40, end_row=7, end_column=45)
                ws.merge_cells(start_row=18, start_column=6, end_row=18, end_column=11)
                ws.merge_cells(start_row=19, start_column=26, end_row=19, end_column=31)
                ws.merge_cells(start_row=19, start_column=40, end_row=19, end_column=45)
                ws.merge_cells(start_row=30, start_column=6, end_row=30, end_column=11)
                ws.merge_cells(start_row=31, start_column=26, end_row=31, end_column=31)
                ws.merge_cells(start_row=31, start_column=40, end_row=31, end_column=45)

                for y in range(9, 15):
                    ws.merge_cells(start_row=y, start_column=3, end_row=y, end_column=12) # col 1
                    ws.merge_cells(start_row=y, start_column=13, end_row=y, end_column=24) # col 2

                for y in range (21, 27):
                    ws.merge_cells(start_row=y, start_column=3, end_row=y, end_column=12) # col 1
                    ws.merge_cells(start_row=y, start_column=13, end_row=y, end_column=24) # col 2

                for y in range (33, 39):
                    ws.merge_cells(start_row=y, start_column=3, end_row=y, end_column=12) # col 1
                    ws.merge_cells(start_row=y, start_column=13, end_row=y, end_column=24) # col 2

                for y in range(9, 15):
                    for x in range(3, 25):
                        ws.cell(row=y, column=x).border = Border(top=thin, left=thin, bottom=thin, right=thin)

                for y in range(21, 27):
                    for x in range(3, 25):
                        ws.cell(row=y, column=x).border = Border(top=thin, left=thin, bottom=thin, right=thin)

                for y in range(33, 39):
                    for x in range(3, 25):
                        ws.cell(row=y, column=x).border = Border(top=thin, left=thin, bottom=thin, right=thin)

                table_row_num = 0
                first_row = True
                for row in rows:
                    print(row)
                    table_row_num += 1
                    process_type = row[0]
                    temperature = row[1]
                    temperature_unit = row[2]
                    heater_set = row[3]
                    heater_set_unit = row[4]
                    timestamp = row[5].strftime("%H:%M:%S")
                    if first_row:
                        prev_process_type = process_type
                        first_row = False
                    
                    if prev_process_type != process_type:
                        table_row_num = 1

                    if process_type == 'post production day 1':
                        table_first_row = 9
                    elif process_type == 'post production day 2':
                        table_first_row = 21
                    elif process_type == 'post production day 3':
                        table_first_row = 33

                    if table_row_num <=5:
                        ws.cell(row=table_first_row+table_row_num, column=3).value = timestamp
                        ws.cell(row=table_first_row+table_row_num, column=3).font = font_size_10
                        ws.cell(row=table_first_row+table_row_num, column=3).alignment = Alignment(horizontal='center')
                        ws.cell(row=table_first_row+table_row_num, column=13).value = '%s %s' % (temperature, temperature_unit)
                        ws.cell(row=table_first_row+table_row_num, column=13).font = font_size_10
                        ws.cell(row=table_first_row+table_row_num, column=13).alignment = Alignment(horizontal='center')
                    prev_process_type = process_type

                report_name = "post_production_%s" % (process_date.strftime("%-d%b%Y"))
                
            else:
                valid_data = False
                log.error('invalid log data')

        if valid_data:
            file_path = Path("/home/pi/monitoring_sys/core/static/assets/reports/%s.xlsx" % report_name)
            while file_path.is_file():
                log.debug('Report file name exist')
                report_name += '(1)'
                file_path = Path("/home/pi/monitoring_sys/core/static/assets/reports/%s.xlsx" % report_name)
            wb.save(file_path)
            log.info('Report generated: %s' % report_name)
            database.query('local', 'update', 'UPDATE app_report SET file_name="%s" WHERE id=%s' % (report_name, batch_id))
            return 0
    else:
        log.error('Trying to generate report but no report data found.')
    return 1




if __name__ == '__main__':
    log = logger.init(os.path.abspath(__file__), '') # logging
    database.init(log) # database
    # generate(log, 89) # pre production
    # generate(log, 91) # production
    generate(log, 95) # post production